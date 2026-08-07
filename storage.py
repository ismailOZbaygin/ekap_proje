"""CSV / XLSX storage layer — Dictionary-based Upsert Architecture.

Implements a full in-memory dictionary storage approach.
Reads the entire CSV into RAM, updates records by IKN (O(1)), 
and rewrites the entire file upon changes.

Warning: This is fine for < 100,000 records. If your dataset grows 
beyond that, migrate to SQLite.
"""

from __future__ import annotations

import csv
import logging
import os
from typing import Optional

from models import ContractRecord
import config

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════
#  Phase 1 — State Hydration (In-Memory Dictionary)
# ═══════════════════════════════════════════════════════════════════════

class IKNCache:
    """Hash-map of records already persisted in the CSV.

    Keys are IKNs, values are dictionaries representing the row.
    This allows O(1) lookups and O(1) in-memory updates.
    """

    def __init__(self) -> None:
        self._data: dict[str, dict] = {}

    def load(self, filepath: Optional[str] = None) -> int:
        """Read existing CSV and populate the dictionary."""
        path = filepath or _csv_path()
        if not os.path.isfile(path):
            logger.info("CSV dosyası bulunamadı — boş cache ile başlanıyor.")
            return 0

        count = 0
        try:
            with open(path, "r", encoding=config.CSV_ENCODING, newline="") as f:
                content = f.read()
                if not content.strip():
                    return 0

                # Otomatik ayırıcı (delimiter) tespiti (virgül vs noktalı virgül)
                first_line = content.splitlines()[0]
                delimiter = config.CSV_DELIMITER
                if ";" not in first_line and "," in first_line:
                    delimiter = ","

            with open(path, "r", encoding=config.CSV_ENCODING, newline="") as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                for row in reader:
                    ikn = row.get("İKN")
                    if ikn:
                        self._data[ikn] = row
                        count += 1
        except Exception as exc:
            logger.warning("CSV okuma hatası (cache): %s", exc)

        logger.info("Cache yüklendi — %d kayıt bellekte.", count)
        return count

    def get_all_records(self) -> list[dict]:
        """Return all current records as a list of dicts."""
        return list(self._data.values())

    def update_record(self, record: ContractRecord) -> bool:
        """Update or insert a record in memory.
        
        Returns True if it was a new insert, False if it was an update.
        """
        ikn = record.ikn
        is_new = ikn not in self._data
        
        new_data = record.to_dict()
        if not is_new:
            current_data = self._data[ikn]
            for key, value in new_data.items():
                if value: # Yeni veri boş değilse eskisinin üstüne yaz
                    current_data[key] = value
            self._data[ikn] = current_data
            logger.debug("Bellekte güncellendi: %s", ikn)
        else:
            self._data[ikn] = new_data
            logger.debug("Belleğe eklendi: %s", ikn)
            
        return is_new

    def __len__(self) -> int:
        return len(self._data)


# ═══════════════════════════════════════════════════════════════════════
#  Phase 4 — I/O (Full Rewrite)
# ═══════════════════════════════════════════════════════════════════════

def _ensure_output_dir() -> None:
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)


def _csv_path() -> str:
    return os.path.join(config.OUTPUT_DIR, config.OUTPUT_FILE)


def _rewrite_entire_csv(cache: IKNCache) -> None:
    """Writes the entire in-memory dictionary back to the CSV."""
    _ensure_output_dir()
    path = _csv_path()
    
    records = cache.get_all_records()
    if not records:
        return

    with open(path, "w", encoding=config.CSV_ENCODING, newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=ContractRecord.csv_headers(),
            delimiter=config.CSV_DELIMITER,
        )
        writer.writeheader()
        writer.writerows(records)


def append_record(record: ContractRecord, cache: IKNCache) -> bool:
    """Upsert a single record and rewrite CSV."""
    is_new = cache.update_record(record)
    _rewrite_entire_csv(cache)
    
    if getattr(config, "EXPORT_XLSX", True):
        try:
            export_to_xlsx()
        except Exception as exc:
            logger.debug("Auto XLSX export hatası: %s", exc)

    if is_new:
        logger.info("Yeni kayıt yazıldı: %s", record.ikn)
    else:
        logger.info("Mevcut kayıt güncellendi: %s", record.ikn)
        
    return is_new


def append_batch(records: list[ContractRecord], cache: IKNCache) -> int:
    """Upsert a batch of records and rewrite CSV ONCE at the end.
    
    Returns the number of strictly NEW records added.
    """
    if not records:
        return 0

    new_count = 0
    update_count = 0
    
    for rec in records:
        is_new = cache.update_record(rec)
        if is_new:
            new_count += 1
        else:
            update_count += 1

    # O(N) maliyetli yazma işlemini döngü dışında sadece 1 kez yapıyoruz
    _rewrite_entire_csv(cache)

    if getattr(config, "EXPORT_XLSX", True):
        try:
            export_to_xlsx()
        except Exception as exc:
            logger.debug("Auto XLSX export hatası: %s", exc)

    logger.info("Batch Upsert tamamlandı — %d yeni eklendi, %d güncellendi.", new_count, update_count)
    return new_count


# ═══════════════════════════════════════════════════════════════════════
#  Optional XLSX Export
# ═══════════════════════════════════════════════════════════════════════

def export_to_xlsx(xlsx_filename: Optional[str] = None) -> str:
    """Convert the current CSV to a styled XLSX file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.warning("openpyxl yüklü değil, XLSX oluşturulamadı.")
        return ""

    csv_path = _csv_path()
    if not os.path.isfile(csv_path):
        return ""

    xlsx_name = xlsx_filename or config.OUTPUT_FILE.replace(".csv", ".xlsx")
    xlsx_path = os.path.join(config.OUTPUT_DIR, xlsx_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sözleşme Verileri"

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    with open(csv_path, "r", encoding=config.CSV_ENCODING, newline="") as f:
        reader = csv.reader(f, delimiter=config.CSV_DELIMITER)
        for row_idx, row in enumerate(reader, 1):
            ws.append(row)
            for col_idx in range(1, len(row) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = cell_font

    ws.row_dimensions[1].height = 25

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=12)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(xlsx_path)
    logger.info("XLSX oluşturuldu: %s", xlsx_path)
    return xlsx_path